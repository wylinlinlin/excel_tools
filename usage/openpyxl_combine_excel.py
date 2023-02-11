#!/usr/bin/python3
# --*-- coding: utf-8 --*--
 
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
 
 
def form_merge():
    """文件合并"""
    result_path = "C:\\Users\\hanstan\\Desktop\\485496\\科目余额表\\21\\"  # 获取目标文件夹
    file_list, start_row = [], 0
    for i, j, k in os.walk(result_path):  # 1、遍历获取目标文件夹下的所有xlsx文件
        file_list = k
    file_list = filter(lambda x: 'xlsx' in x, file_list)  # 过滤掉不是xlsx格式的文件
    lb = load_workbook('test.xlsx')
    sheet = lb['sheet1']   # 2、打开/新建一个xlsx总表用于写入(合并)所有表数据
    for file in file_list:
        lb1 = load_workbook(result_path + file)
        print(f"read {file}...")
        sheet1 = lb1['sheet1']  # 3、打开每个目标xlsx表，开始写入数据
        print(f"{str(file)} is writed...\n")
        for row in range(1, sheet1.max_row + 1):
            for column in range(1, sheet1.max_column + 1):
                value = sheet1.cell(row, column).value  # 数据的文本
                
                if row == 1 and start_row != 0:  # 6、需要对后续表的首行作判断，用于删除多余的表头
                    row += 1
                    start_row -= 1
                
                sheet.cell(row + start_row, column, value)
        start_row += sheet1.max_row  # 5、当前表写完数据后记住最后的行数，便于下个表续写追加内容
    lb.save('C:\\Users\\hanstan\\Desktop\\485496\\TB21.xlsx')  # 7、保存表
    print('over')
 
 
 
 
if __name__ == '__main__':
    form_merge()

 